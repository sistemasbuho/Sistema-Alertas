from contextlib import nullcontext
from datetime import date, datetime, time
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import SimpleTestCase
from django.utils import timezone
from django.utils.datastructures import MultiValueDict
from rest_framework.response import Response
from rest_framework.test import APIRequestFactory

from apps.base.api.ingestion import IngestionAPIView
from apps.base.api.utils import formatear_fecha_respuesta
from apps.base.models import Articulo


class IngestionAPITests(SimpleTestCase):
    def setUp(self):
        self.factory = APIRequestFactory()
        self.proyecto_id = "123e4567-e89b-12d3-a456-426614174000"
        self.ultimo_registro_articulo = None
        self.ultimo_registro_red = None

    def _mock_proyecto(
        self, mock_proyecto, criterios=None, tipo_alerta="medios", nombre="Proyecto Test"
    ):
        criterios = criterios or []
        proyecto = SimpleNamespace(
            id=self.proyecto_id,
            get_criterios_aceptacion_list=lambda: criterios,
            tipo_alerta=tipo_alerta,
            nombre=nombre,
        )
        mock_proyecto.objects.filter.return_value.first.return_value = proyecto

    @patch("apps.base.api.ingestion.Proyecto")
    def test_detects_medios_twk_from_csv_and_forwards_payload(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto)
        content = (
            "title,content,published,extra_source_attributes.name,reach,url\n"
            "Titulo,Contenido,2024-01-01,Fuente,1000,http://example.com/articulo\n"
        )
        uploaded = SimpleUploadedFile(
            "medios.csv",
            content.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        listado = response.data["listado"]
        self.assertEqual(len(listado), 1)
        self.assertEqual(listado[0]["autor"], "Fuente")
        self.assertEqual(listado[0]["tipo"], "medios")
        self.assertEqual(response.data["duplicados"], 0)
        self.assertEqual(response.data["descartados"], 0)
        self.assertEqual(response.data["mensaje"], "1 registros creados")
        self.assertEqual(response.data["proyecto_nombre"], "Proyecto Test")

    @patch("apps.base.api.ingestion.Proyecto")
    @patch("apps.base.api.ingestion.enviar_alertas_automatico")
    def test_envio_automatico_envia_alertas(self, mock_enviar, mock_proyecto):
        self._mock_proyecto(mock_proyecto)
        proyecto = mock_proyecto.objects.filter.return_value.first.return_value
        proyecto.tipo_envio = "automatico"
        content = (
            "title,content,published,extra_source_attributes.name,reach,url\n"
            "Titulo,Contenido,2024-01-01,Fuente,1000,http://example.com/articulo\n"
        )
        uploaded = SimpleUploadedFile(
            "medios.csv",
            content.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        mock_enviar.return_value = {"success": True}

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ), patch.object(
            IngestionAPIView,
            "_notificar_ruta_externa",
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        mock_enviar.assert_called_once()
        args, kwargs = mock_enviar.call_args
        self.assertEqual(args[0], proyecto.id)
        self.assertEqual(args[1], "medios")
        self.assertEqual(len(args[2]), 1)
        self.assertEqual(args[2][0]["id"], "articulo-id")
        self.assertEqual(kwargs.get("usuario_id"), 2)

    @patch("apps.base.api.ingestion.Proyecto")
    @patch("apps.base.api.ingestion.enviar_alertas_automatico")
    def test_envio_manual_no_envia_alertas(self, mock_enviar, mock_proyecto):
        self._mock_proyecto(mock_proyecto)
        proyecto = mock_proyecto.objects.filter.return_value.first.return_value
        proyecto.tipo_envio = "manual"
        content = (
            "title,content,published,extra_source_attributes.name,reach,url\n"
            "Titulo,Contenido,2024-01-01,Fuente,1000,http://example.com/articulo\n"
        )
        uploaded = SimpleUploadedFile(
            "medios.csv",
            content.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ), patch.object(
            IngestionAPIView,
            "_notificar_ruta_externa",
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        mock_enviar.assert_not_called()

    @patch("apps.base.api.ingestion.Proyecto")
    def test_detects_global_news_provider(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto)
        content = (
            "Título,Resumen - Aclaracion,Fecha,Autor - Conductor,Medio,URL\n"
            "Noticia,Contenido GN,2024-03-01,Autor GN,Canal GN,http://example.com/global-news\n"
        )
        uploaded = SimpleUploadedFile(
            "global_news.csv",
            content.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["proveedor"], "global_news")
        listado = response.data["listado"]
        self.assertEqual(len(listado), 1)
        alerta = listado[0]
        self.assertEqual(alerta["autor"], "Canal GN")
        self.assertEqual(alerta["titulo"], "Noticia")
        self.assertEqual(alerta["contenido"], "Contenido GN")

    @patch.object(Articulo.objects, "filter")
    def test_es_url_duplicada_por_proyecto_normaliza_variantes(self, mock_filter):
        view = IngestionAPIView()
        proyecto = SimpleNamespace(id=self.proyecto_id)

        mock_queryset = mock_filter.return_value
        mock_queryset.values_list.return_value = ["https://example.com/noticia"]

        es_duplicado = view._es_url_duplicada_por_proyecto(
            Articulo,
            proyecto,
            "http://www.example.com/noticia/",
        )

        self.assertTrue(es_duplicado)
        mock_filter.assert_called_once_with(proyecto=proyecto)

    @patch("apps.base.api.ingestion.Proyecto")
    def test_detects_stakeholders_provider(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto)
        content = (
            "Titular,Resumen,Fecha,Autor,Fuente,URL\n"
            "Stakeholder News,Resumen SH,2024-04-15,Autor SH,FUENTE,http://example.com/stakeholders\n"
        )
        uploaded = SimpleUploadedFile(
            "stakeholders.csv",
            content.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["proveedor"], "stakeholders")
        listado = response.data["listado"]
        self.assertEqual(len(listado), 1)
        alerta = listado[0]
        self.assertEqual(alerta["autor"], "FUENTE")
        self.assertEqual(alerta["titulo"], "Stakeholder News")
        self.assertEqual(alerta["contenido"], "Resumen SH")

    @patch("apps.base.api.ingestion.Proyecto")
    def test_stakeholders_no_combina_fecha_y_hora(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto)
        content = (
            "Titular,Resumen,Fecha,Hora,Autor,Fuente,URL\n"
            "Stakeholder News,Resumen SH,2024-04-15,08:30,Autor SH,FUENTE,http://example.com/stakeholders\n"
        )
        uploaded = SimpleUploadedFile(
            "stakeholders.csv",
            content.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        self.assertIsNotNone(self.ultimo_registro_articulo)
        fecha_resultado = self.ultimo_registro_articulo.get("fecha")
        self.assertIsNotNone(fecha_resultado)
        self.assertTrue(timezone.is_aware(fecha_resultado))
        self.assertEqual(fecha_resultado.date(), date(2024, 4, 15))
        self.assertEqual(fecha_resultado.hour, 0)
        self.assertEqual(fecha_resultado.minute, 0)

    @patch("apps.base.api.ingestion.Proyecto")
    def test_detects_determ_medios_provider(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto)
        content = (
            "TITLE,MENTION_SNIPPET,DATE,TIME,REACH,ENGAGEMENT_RATE,AUTHOR,FROM,URL\n"
            "Determinacion,Resumen DM,2024-05-20,08:15,2000,7,Autor DM,Fuente DM,http://example.com/determ-medios\n"
        )
        uploaded = SimpleUploadedFile(
            "determ_medios.csv",
            content.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["proveedor"], "determ_medios")
        listado = response.data["listado"]
        self.assertEqual(len(listado), 1)
        alerta = listado[0]
        self.assertEqual(alerta["autor"], "Fuente DM")
        self.assertEqual(alerta["titulo"], "Determinacion")
        self.assertEqual(alerta["contenido"], "Resumen DM")

    @patch("apps.base.api.ingestion.Proyecto")
    def test_determ_prefiere_columna_author_en_redes(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto, tipo_alerta="redes")
        content = (
            "mention_snippet,date,time,reach,engagement_rate,AUTHOR,FROM,url,social_network\n"
            "Contenido determin,2024-05-21,10:00,1500,12,Autor Preferido,Autor Alterno,http://example.com/determ,Twitter\n"
        )
        uploaded = SimpleUploadedFile(
            "determ_redes.csv",
            content.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        self.assertIsNotNone(self.ultimo_registro_red)
        self.assertEqual(self.ultimo_registro_red.get("autor"), "Autor Preferido")
        self.assertEqual(response.data["listado"][0]["autor"], "Autor Preferido")

    @patch("apps.base.api.ingestion.Proyecto")
    def test_rechaza_archivo_sin_columna_url(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto)
        content = (
            "title,content,published,extra_source_attributes.name,reach\n"
            "Titulo,Contenido,2024-01-01,Fuente,1000\n"
        )
        uploaded = SimpleUploadedFile(
            "medios.csv",
            content.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 400)
        self.assertIn("columna 'url'", response.data["detail"].lower())

    @patch("apps.base.api.ingestion.Proyecto")
    def test_rechaza_archivo_con_columna_url_sin_datos(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto)
        content = (
            "title,content,published,extra_source_attributes.name,reach,url\n"
            "Titulo,Contenido,2024-01-01,Fuente,1000,\n"
        )
        uploaded = SimpleUploadedFile(
            "medios.csv",
            content.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 400)
        self.assertIn("valor válido", response.data["detail"].lower())

    def test_parse_xlsx_ignora_columnas_vacias(self):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "title",
                "content",
                "published",
                "extra_source_attributes.name",
                "reach",
                "url",
                "columna_vacia",
            ]
        )
        sheet.append(
            [
                "Titulo",
                "Contenido",
                "2024-01-01",
                "Fuente",
                "1000",
                "http://example.com/desde-xlsx",
                None,
            ]
        )
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        view = IngestionAPIView()
        headers, rows = view._parse_xlsx(buffer)

        self.assertIn("url", headers)
        self.assertTrue(rows)
        self.assertNotIn("columna_vacia", rows[0])

    def test_normalizar_columnas_url_usa_link_como_url(self):
        view = IngestionAPIView()
        headers = ["titulo", "link"]
        rows = [{"titulo": "Noticia", "link": "http://example.com/link"}]

        nuevos_headers, nuevos_rows = view._normalizar_columnas_url(headers, rows)

        self.assertIn("url", nuevos_headers)
        self.assertEqual(nuevos_rows[0]["url"], "http://example.com/link")

    def test_parse_xlsx_extrae_hipervinculo_en_columna_link_streaming(self):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Link (Streaming – Imagen)", "Titulo"])
        cell = sheet.cell(row=2, column=1, value="Ver")
        cell.hyperlink = "http://example.com/stream"
        sheet.cell(row=2, column=2, value="Titulo Global")

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        view = IngestionAPIView()
        headers, rows = view._parse_xlsx(buffer)
        headers, rows = view._normalizar_columnas_url(headers, rows)

        self.assertIn("url", headers)
        self.assertTrue(rows)
        self.assertEqual(rows[0]["url"], "http://example.com/stream")

    def test_mapear_medios_global_news_prioriza_fecha_y_link_streaming(self):
        view = IngestionAPIView()
        row = {
            "titulo": "Titulo GN",
            "resumen - aclaracion": "Contenido GN",
            "fecha": "2024-05-10",
            "published": "2024-05-01",
            "autor - conductor": "Autor GN",
            "medio": "Canal GN",
            "link (streaming – imagen)": "http://example.com/streaming",
            "audiencia": "2500",
        }

        registro = view._mapear_medios_twk(row, "global_news")

        self.assertEqual(registro["url"], "http://example.com/streaming")
        self.assertIsNotNone(registro["fecha"])
        self.assertEqual(registro["fecha"].date(), date(2024, 5, 10))
        self.assertEqual(registro["reach"], 2500)
        self.assertEqual(registro["autor"], "Canal GN")

    def test_mapear_medios_global_news_soporta_formato_fecha_latam(self):
        view = IngestionAPIView()
        row = {
            "titulo": "Titulo GN",
            "contenido": "Contenido GN",
            "fecha": "11/09/2025",
            "autor": "Autor GN",
            "medio": "Canal GN",
            "audiencia": "1500",
            "url": "http://example.com/global-news",
        }

        registro = view._mapear_medios_twk(row, "global_news")

        self.assertIsNotNone(registro["fecha"])
        self.assertEqual(registro["fecha"].date(), date(2025, 9, 11))
        self.assertEqual(registro["reach"], 1500)
        self.assertEqual(registro["autor"], "Canal GN")

    def test_mapear_medios_stakeholders_no_combina_fecha_iso_y_hora(self):
        view = IngestionAPIView()
        row = {
            "titulo": "Titulo Stakeholder",
            "resumen": "Contenido",
            "fecha": "2025-09-29",
            "hora": "08:30",
            "autor": "Autor SH",
            "fuente": "Fuente Stakeholder",
            "audiencia": "200",
            "url": "http://example.com/stakeholder",
        }

        registro = view._mapear_medios_twk(row, "stakeholders")

        self.assertIsNotNone(registro["fecha"])
        self.assertEqual(registro["fecha"].date(), date(2025, 9, 29))
        self.assertEqual(registro["fecha"].time(), time.min)
        self.assertEqual(registro["autor"], "Fuente Stakeholder")

    @patch("apps.base.api.ingestion.Proyecto")
    def test_respuesta_incluye_conteo_de_duplicados(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto)
        content = (
            "title,content,published,extra_source_attributes.name,reach,url\n"
            "Titulo,Contenido,2024-01-01,Fuente,1000,http://example.com/articulo\n"
        )
        uploaded = SimpleUploadedFile(
            "medios.csv",
            content.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=ValueError("La URL ya existe para este proyecto"),
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["duplicados"], 1)
        self.assertEqual(response.data["descartados"], 0)
        self.assertIn("1 duplicados", response.data["mensaje"])
        self.assertEqual(len(response.data["errores"]), 1)

    @patch("apps.base.api.ingestion.Proyecto")
    def test_respuesta_sin_registros_incluye_nombre(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto)
        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}", {}, format="multipart"
        )

        with patch.object(
            IngestionAPIView,
            "_extraer_registros_estandar",
            return_value=([{"proveedor": "medios"}], "medios", None),
        ), patch.object(
            IngestionAPIView, "_filtrar_por_criterios", return_value=[]
        ), patch.object(
            IngestionAPIView, "_notificar_ruta_externa"
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 405)
        self.assertEqual(response.data["proyecto_nombre"], "Proyecto Test")

    @patch("apps.base.api.ingestion.Proyecto")
    def test_detects_redes_twk_from_xlsx_and_forwards_payload(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto, tipo_alerta="redes")
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "content",
                "published",
                "extra_author_attributes.name",
                "reach",
                "engagement",
                "url",
                "red_social",
            ]
        )
        sheet.append(
            [
                "Hola",
                "2024-02-02",
                "User",
                "500",
                "42",
                "http://example.com/red-uno",
                "Twitter",
            ]
        )
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        uploaded = SimpleUploadedFile(
            "redes.xlsx",
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        listado = response.data["listado"]
        self.assertEqual(len(listado), 1)
        alerta = listado[0]
        self.assertIsNone(self.ultimo_registro_articulo)
        self.assertIsNotNone(self.ultimo_registro_red)
        self.assertEqual(alerta["contenido"], self.ultimo_registro_red.get("contenido"))
        self.assertEqual(alerta["engagement"], self.ultimo_registro_red.get("engagement"))
        self.assertEqual(alerta["tipo"], "redes")
        self.assertEqual(response.data["duplicados"], 0)
        self.assertEqual(response.data["descartados"], 0)
        self.assertEqual(response.data["mensaje"], "1 registros creados")

    @patch("apps.base.api.ingestion.Proyecto")
    def test_redes_twk_prefiere_short_name_para_autor(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto, tipo_alerta="redes")
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "content",
                "published",
                "extra_author_attributes.short_name",
                "reach",
                "engagement",
                "url",
                "red_social",
            ]
        )
        sheet.append(
            [
                "Hola desde redes",
                "2024-02-03",
                "Alias corto",
                "1200",
                "84",
                "http://example.com/red-short",
                "Twitter",
            ]
        )
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        uploaded = SimpleUploadedFile(
            "redes.xlsx",
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        self.assertIsNotNone(self.ultimo_registro_red)
        self.assertEqual(self.ultimo_registro_red.get("autor"), "Alias corto")
        self.assertEqual(response.data["listado"][0]["autor"], "Alias corto")

    @patch("apps.base.api.ingestion.Proyecto")
    def test_respuesta_incluye_conteo_de_descartados(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto)
        content = (
            "title,content,published,extra_source_attributes.name,reach,url\n"
            "Titulo,Contenido,2024-01-01,Fuente,1000,http://example.com/articulo\n"
        )
        uploaded = SimpleUploadedFile(
            "medios.csv",
            content.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=RuntimeError("Error inesperado"),
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["duplicados"], 0)
        self.assertEqual(response.data["descartados"], 1)
        self.assertIn("1 descartados", response.data["mensaje"])
        self.assertEqual(len(response.data["errores"]), 1)

    @patch("apps.base.api.ingestion.Proyecto")
    def test_redes_twk_trim_contenido_for_twitter_qt(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto, tipo_alerta="redes")
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "content",
                "published",
                "extra_author_attributes.name",
                "reach",
                "engagement",
                "url",
                "red_social",
            ]
        )
        sheet.append(
            [
                "Mensaje inicial QT @usuario comentario",
                "2024-03-03",
                "User",
                "500",
                "42",
                "http://example.com/red-dos",
                "Twitter",
            ]
        )
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        uploaded = SimpleUploadedFile(
            "redes.xlsx",
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        alerta = response.data["listado"][0]
        self.assertIsNotNone(self.ultimo_registro_red)
        self.assertEqual(alerta["contenido"], self.ultimo_registro_red.get("contenido"))
        self.assertEqual(alerta["red_social"], None)
        self.assertEqual(alerta["tipo"], "redes")
        self.assertEqual(response.data["duplicados"], 0)
        self.assertEqual(response.data["descartados"], 0)

    @patch("apps.base.api.ingestion.Proyecto")
    def test_permite_multiples_archivos_en_una_misma_peticion(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto)
        content_uno = (
            "title,content,published,extra_source_attributes.name,reach,url\n"
            "Titulo 1,Contenido 1,2024-01-01,Fuente 1,1000,http://example.com/articulo-1\n"
        )
        content_dos = (
            "title,content,published,extra_source_attributes.name,reach,url\n"
            "Titulo 2,Contenido 2,2024-02-02,Fuente 2,500,http://example.com/articulo-2\n"
        )
        uploaded_uno = SimpleUploadedFile(
            "medios-uno.csv",
            content_uno.encode("utf-8"),
            content_type="text/csv",
        )
        uploaded_dos = SimpleUploadedFile(
            "medios-dos.csv",
            content_dos.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": [uploaded_uno, uploaded_dos]},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        listado = response.data["listado"]
        self.assertEqual(len(listado), 2)
        autores = {alerta["autor"] for alerta in listado}
        self.assertSetEqual(autores, {"Fuente 1", "Fuente 2"})
        self.assertEqual(response.data["duplicados"], 0)
        self.assertEqual(response.data["descartados"], 0)
        self.assertEqual(response.data["mensaje"], "2 registros creados")

    @patch("apps.base.api.ingestion.Proyecto")
    def test_permite_archivos_en_claves_distintas(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto)
        contenido_a = (
            "title,content,published,extra_source_attributes.name,reach,url\n"
            "Titulo A,Contenido A,2024-03-03,Fuente A,1500,http://example.com/articulo-a\n"
        )
        contenido_b = (
            "title,content,published,extra_source_attributes.name,reach,url\n"
            "Titulo B,Contenido B,2024-04-04,Fuente B,800,http://example.com/articulo-b\n"
        )
        archivo_a = SimpleUploadedFile(
            "medios-a.csv",
            contenido_a.encode("utf-8"),
            content_type="text/csv",
        )
        archivo_b = SimpleUploadedFile(
            "medios-b.csv",
            contenido_b.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": archivo_a, "archivo_secundario": archivo_b},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        listado = response.data["listado"]
        self.assertEqual(len(listado), 2)
        autores = {alerta["autor"] for alerta in listado}
        self.assertSetEqual(autores, {"Fuente A", "Fuente B"})
        self.assertEqual(response.data["duplicados"], 0)
        self.assertEqual(response.data["descartados"], 0)
        self.assertEqual(response.data["mensaje"], "2 registros creados")

    @patch("apps.base.api.ingestion.Proyecto")
    def test_aplica_filtro_de_criterios_de_aceptacion(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto, criterios=["alerta"])
        content = (
            "title,content,published,extra_source_attributes.name,reach,url\n"
            "Mensaje sin match,Contenido,2024-01-01,Fuente,1000,http://example.com/articulo-sin-match\n"
            "Alerta importante,Contenido,2024-01-01,Fuente,1000,http://example.com/articulo-alerta\n"
        )
        uploaded = SimpleUploadedFile(
            "medios.csv",
            content.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        payload = response.data
        self.assertEqual(len(payload["listado"]), 1)
        self.assertEqual(payload["listado"][0]["titulo"], "Alerta importante")
        self.assertEqual(payload["duplicados"], 0)
        self.assertEqual(payload["descartados"], 0)
        self.assertEqual(payload["mensaje"], "1 registros creados")

    @patch("apps.base.api.ingestion.Proyecto")
    def test_filtro_de_criterios_sin_coincidencias_no_reenvia(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto, criterios=["alerta"])
        content = (
            "title,content,published,extra_author_attributes.name,reach,url\n"
            "Mensaje sin match,Contenido,2024-01-01,Autor,1000,http://example.com/articulo-filtrado\n"
        )
        uploaded = SimpleUploadedFile(
            "medios.csv",
            content.encode("utf-8"),
            content_type="text/csv",
        )

        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {"archivo": uploaded},
            format="multipart",
        )

        with patch.object(
            IngestionAPIView,
            "forward_payload",
            return_value=Response({"ok": True}, status=202),
        ) as mock_forward:
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 405)
        mock_forward.assert_not_called()
        self.assertIn("mensaje", response.data)
        self.assertIn("criterios", response.data["mensaje"])
        self.assertEqual(response.data["duplicados"], 0)
        self.assertEqual(response.data["descartados"], 0)

    def test_construir_payload_forward_usa_tipo_alerta_del_proyecto(self):
        proyecto = SimpleNamespace(id=self.proyecto_id, tipo_alerta="redes")
        registros = [
            {
                "tipo": "articulo",
                "titulo": "Titulo",
                "contenido": "Contenido",
                "fecha": datetime(2024, 1, 1),
                "autor": "Autor",
                "reach": 100,
                "engagement": 5,
                "url": "http://example.com",
                "red_social": "Twitter",
                "datos_adicionales": {"extra": "valor"},
                "proveedor": "medios_twk",
            }
        ]

        view = IngestionAPIView()
        payload = view._construir_payload_forward("medios", registros, proyecto)

        self.assertEqual(payload["proveedor"], "medios_twk")
        self.assertEqual(payload["alertas"][0]["tipo"], "redes")

    @patch("apps.base.api.ingestion.Proyecto")
    def test_registro_manual_usa_tipo_alerta_redes_del_proyecto(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto, tipo_alerta="redes")
        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {
                "proyecto": self.proyecto_id,
                "url": "http://example.com/post",
                "contenido": "Contenido redes",
                "autor": "Usuario",
                "red_social": "https://twitter.com", 
            },
            format="json",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        self.assertIsNone(self.ultimo_registro_articulo)
        self.assertIsNotNone(self.ultimo_registro_red)
        self.assertEqual(response.data["listado"][0]["tipo"], "redes")
        self.assertEqual(response.data["duplicados"], 0)
        self.assertEqual(response.data["descartados"], 0)
        self.assertEqual(response.data["mensaje"], "1 registros creados")

    @patch("apps.base.api.ingestion.Proyecto")
    def test_registro_manual_usa_tipo_alerta_medios_del_proyecto(self, mock_proyecto):
        self._mock_proyecto(mock_proyecto, tipo_alerta="medios")
        request = self.factory.post(
            f"/api/ingestion/?proyecto={self.proyecto_id}",
            {
                "proyecto": self.proyecto_id,
                "url": "http://example.com/articulo",
                "titulo": "Titulo articulo",
                "contenido": "Contenido articulo",
                "autor": "Usuario",
                "tipo": "red",
            },
            format="json",
        )

        with patch.object(
            IngestionAPIView,
            "_obtener_usuario_sistema",
            return_value=SimpleNamespace(id=2),
        ), patch.object(
            IngestionAPIView,
            "_crear_articulo",
            side_effect=self._fake_crear_articulo,
        ), patch.object(
            IngestionAPIView,
            "_crear_red_social",
            side_effect=self._fake_crear_red_social,
        ):
            response = IngestionAPIView.as_view()(request)

        self.assertEqual(response.status_code, 201)
        self.assertIsNotNone(self.ultimo_registro_articulo)
        self.assertIsNone(self.ultimo_registro_red)
        self.assertEqual(response.data["listado"][0]["tipo"], "medios")
        self.assertEqual(response.data["duplicados"], 0)
        self.assertEqual(response.data["descartados"], 0)
        self.assertEqual(response.data["mensaje"], "1 registros creados")

    def _fake_crear_articulo(self, registro, proyecto, sistema_user):
        self.ultimo_registro_articulo = registro
        return SimpleNamespace(
            id="articulo-id",
            titulo=registro.get("titulo"),
            contenido=registro.get("contenido"),
            fecha_publicacion=registro.get("fecha"),
            autor=registro.get("autor"),
            reach=registro.get("reach"),
            url=registro.get("url"),
        )

    def _fake_crear_red_social(self, registro, proyecto):
        self.ultimo_registro_red = registro
        red_social_nombre = registro.get("red_social")
        red_social = (
            SimpleNamespace(nombre=red_social_nombre)
            if red_social_nombre
            else None
        )
        return SimpleNamespace(
            id="red-id",
            contenido=registro.get("contenido"),
            fecha_publicacion=registro.get("fecha"),
            autor=registro.get("autor"),
            reach=registro.get("reach"),
            engagement=registro.get("engagement"),
            url=registro.get("url"),
            red_social=red_social,
        )


class IngestionPersistenceTests(SimpleTestCase):
    def setUp(self):
        self.view = IngestionAPIView()

    @patch.object(IngestionAPIView, "_es_url_duplicada_por_proyecto", return_value=False)
    def test_crear_articulo_determ_medios_crea_detalle_envio(self, _mock_es_url):
        row = {
            "title": "Determinacion",
            "mention_snippet": "Resumen DM",
            "date": "2024-05-20",
            "time": "07:45",
            "reach": "1500",
            "engagement_rate": "12",
            "from": "Fuente DM",
            "author": "Autor DM",
            "url": "http://example.com/determ-medios",
        }

        registro = self.view._mapear_medios_twk(row, "determ_medios")

        self.assertEqual(registro["autor"], "Fuente DM")

        proyecto = SimpleNamespace(id="proyecto-id")
        sistema_user = SimpleNamespace(id=2)
        articulo_creado = SimpleNamespace(
            id="articulo-id",
            titulo=registro.get("titulo"),
            contenido=registro.get("contenido"),
            url=registro.get("url"),
            fecha_publicacion=registro.get("fecha"),
            autor=registro.get("autor"),
            reach=registro.get("reach"),
            proyecto=proyecto,
        )

        with patch(
            "apps.base.api.ingestion.transaction.atomic",
            return_value=nullcontext(),
        ) as mock_atomic, patch(
            "apps.base.api.ingestion.Articulo.objects.create",
            return_value=articulo_creado,
        ) as mock_articulo_create, patch.object(
            IngestionAPIView,
            "_asegurar_detalle_envio",
            return_value=SimpleNamespace(id="detalle-id"),
        ) as mock_detalle_create:
            articulo = self.view._crear_articulo(registro, proyecto, sistema_user)

        self.assertIs(articulo, articulo_creado)
        mock_atomic.assert_called_once()
        mock_articulo_create.assert_called_once()
        mock_detalle_create.assert_called_once_with(
            articulo=articulo_creado,
            proyecto=proyecto,
            usuario=sistema_user,
        )

    @patch.object(IngestionAPIView, "_es_url_duplicada_por_proyecto", return_value=False)
    def test_crear_red_social_crea_detalle_envio(self, _mock_es_url):
        registro = {
            "contenido": "Contenido red",
            "fecha": timezone.now(),
            "url": "http://example.com/red",
            "autor": "Autor Red",
            "reach": 75,
            "engagement": 12,
            "red_social": "Twitter",
        }

        proyecto = SimpleNamespace(id="proyecto-id")
        usuario = SimpleNamespace(id=3)
        red_creada = SimpleNamespace(
            id="red-id",
            contenido=registro.get("contenido"),
            fecha_publicacion=registro.get("fecha"),
            autor=registro.get("autor"),
            reach=registro.get("reach"),
            engagement=registro.get("engagement"),
            url=registro.get("url"),
            red_social=None,
        )

        self.view._usuario_sistema_cache = usuario

        with patch(
            "apps.base.api.ingestion.transaction.atomic",
            return_value=nullcontext(),
        ) as mock_atomic, patch(
            "apps.base.api.ingestion.Redes.objects.create",
            return_value=red_creada,
        ) as mock_red_create, patch.object(
            IngestionAPIView,
            "_asegurar_detalle_envio",
            return_value=SimpleNamespace(id="detalle-id"),
        ) as mock_detalle_create, patch(
            "apps.base.api.ingestion.RedesSociales.objects.filter",
            return_value=SimpleNamespace(first=lambda: None),
        ):
            red = self.view._crear_red_social(registro, proyecto)

        self.assertIs(red, red_creada)
        mock_atomic.assert_called_once()
        mock_red_create.assert_called_once()
        mock_detalle_create.assert_called_once_with(
            red=red_creada,
            proyecto=proyecto,
            usuario=usuario,
        )

    def test_asegurar_detalle_envio_crea_detalle_para_medio(self):
        proyecto = SimpleNamespace(id="proyecto-id")
        articulo = SimpleNamespace(id="articulo-id")
        usuario = SimpleNamespace(id=5)
        detalle_creado = SimpleNamespace(id="detalle-id")

        with patch(
            "apps.base.api.ingestion.DetalleEnvio.objects.get_or_create",
            return_value=(detalle_creado, True),
        ) as mock_get_or_create:
            resultado = self.view._asegurar_detalle_envio(
                articulo=articulo,
                proyecto=proyecto,
                usuario=usuario,
            )

        self.assertIs(resultado, detalle_creado)
        mock_get_or_create.assert_called_once_with(
            medio=articulo,
            proyecto=proyecto,
            defaults={
                "estado_enviado": False,
                "estado_revisado": True,
                "created_by": usuario,
                "modified_by": usuario,
            },
        )

    def test_asegurar_detalle_envio_actualiza_detalle_existente(self):
        proyecto = SimpleNamespace(id="proyecto-id")
        articulo = SimpleNamespace(id="articulo-id")
        usuario = SimpleNamespace(id=6)

        detalle_existente = SimpleNamespace(pk=1)

        def refresh():
            setattr(detalle_existente, "_refrescado", True)

        detalle_existente.refresh_from_db = refresh

        class UpdateCapture:
            def __init__(self):
                self.kwargs = None

            def update(self, **kwargs):
                self.kwargs = kwargs
                return 1

        update_capture = UpdateCapture()

        with patch(
            "apps.base.api.ingestion.DetalleEnvio.objects.get_or_create",
            return_value=(detalle_existente, False),
        ) as mock_get_or_create, patch(
            "apps.base.api.ingestion.DetalleEnvio.objects.filter",
            return_value=update_capture,
        ) as mock_filter:
            resultado = self.view._asegurar_detalle_envio(
                articulo=articulo,
                proyecto=proyecto,
                usuario=usuario,
            )

        self.assertIs(resultado, detalle_existente)
        self.assertTrue(getattr(detalle_existente, "_refrescado", False))
        mock_get_or_create.assert_called_once()
        mock_filter.assert_called_once_with(pk=detalle_existente.pk)
        self.assertEqual(
            update_capture.kwargs,
            {"estado_revisado": True, "modified_by": usuario},
        )

    def test_asegurar_detalle_envio_para_red(self):
        proyecto = SimpleNamespace(id="proyecto-id")
        red = SimpleNamespace(id="red-id")
        usuario = SimpleNamespace(id=7)
        detalle_creado = SimpleNamespace(id="detalle-id")

        with patch(
            "apps.base.api.ingestion.DetalleEnvio.objects.get_or_create",
            return_value=(detalle_creado, True),
        ) as mock_get_or_create:
            resultado = self.view._asegurar_detalle_envio(
                red=red,
                proyecto=proyecto,
                usuario=usuario,
            )

        self.assertIs(resultado, detalle_creado)
        mock_get_or_create.assert_called_once_with(
            red_social=red,
            proyecto=proyecto,
            defaults={
                "estado_enviado": False,
                "estado_revisado": True,
                "created_by": usuario,
                "modified_by": usuario,
            },
        )

    def test_asegurar_detalle_envio_valida_argumentos(self):
        proyecto = SimpleNamespace(id="proyecto-id")
        articulo = SimpleNamespace(id="articulo-id")
        red = SimpleNamespace(id="red-id")

        with self.assertRaises(ValueError):
            self.view._asegurar_detalle_envio(
                proyecto=proyecto,
                usuario=None,
            )

        with self.assertRaises(ValueError):
            self.view._asegurar_detalle_envio(
                proyecto=proyecto,
                usuario=None,
                articulo=articulo,
                red=red,
            )


class ObtenerArchivosTests(SimpleTestCase):
    def setUp(self):
        self.view = IngestionAPIView()

    def test_recupera_todos_los_archivos_desde_multivaluedict(self):
        archivo_uno = SimpleUploadedFile(
            "archivo-uno.csv",
            b"contenido uno",
            content_type="text/csv",
        )
        archivo_dos = SimpleUploadedFile(
            "archivo-dos.csv",
            b"contenido dos",
            content_type="text/csv",
        )
        files = MultiValueDict(
            {
                "archivo_principal": [archivo_uno],
                "archivo_secundario": [archivo_dos],
            }
        )
        request = SimpleNamespace(FILES=files)

        archivos = self.view._obtener_archivos(request)

        self.assertEqual(len(archivos), 2)
        self.assertEqual({a.name for a in archivos}, {"archivo-uno.csv", "archivo-dos.csv"})

    def test_recupera_archivos_desde_diccionario_con_listas(self):
        archivo_uno = SimpleUploadedFile(
            "archivo-uno.csv",
            b"contenido uno",
            content_type="text/csv",
        )
        archivo_dos = SimpleUploadedFile(
            "archivo-dos.csv",
            b"contenido dos",
            content_type="text/csv",
        )
        files = {
            "archivo": [archivo_uno],
            "archivo_extra": archivo_dos,
        }
        request = SimpleNamespace(FILES=files)

        archivos = self.view._obtener_archivos(request)

        self.assertEqual(len(archivos), 2)
        self.assertEqual({a.name for a in archivos}, {"archivo-uno.csv", "archivo-dos.csv"})


class FormatearFechaRespuestaTests(SimpleTestCase):
    def test_formatea_fecha_iso_am(self):
        resultado = formatear_fecha_respuesta("2025-09-23T08:00:38.000Z")
        self.assertEqual(resultado, "2025-09-23 08:00:38 AM")

    def test_formatea_fecha_iso_pm(self):
        resultado = formatear_fecha_respuesta("2025-09-23T16:01:38.000Z")
        self.assertEqual(resultado, "2025-09-23 4:01:38 PM")


class IngestionFormatoFechaTests(SimpleTestCase):
    def setUp(self):
        self.view = IngestionAPIView()
        self.proyecto = SimpleNamespace(
            id="123e4567-e89b-12d3-a456-426614174000",
            tipo_alerta="medios",
        )

    def test_construir_payload_forward_formatea_fecha_iso(self):
        registros = [
            {
                "tipo": "articulo",
                "titulo": "Titulo",
                "contenido": "Contenido",
                "fecha": "2025-09-23T16:01:38.000Z",
                "autor": "Autor",
                "reach": 10,
                "engagement": 5,
                "url": "http://example.com",
                "red_social": None,
                "datos_adicionales": {},
                "proveedor": "medios_twk",
            }
        ]

        payload = self.view._construir_payload_forward("medios", registros, self.proyecto)

        self.assertEqual(payload["alertas"][0]["fecha"], "2025-09-23 4:01:38 PM")

    def test_serializar_articulo_usa_formato_legible(self):
        registro = {
            "proveedor": "medios_twk",
            "datos_adicionales": {},
            "fecha": "2025-09-23T08:00:38.000Z",
        }
        articulo = SimpleNamespace(
            id="articulo-id",
            titulo="Titulo",
            contenido="Contenido",
            fecha_publicacion=None,
            autor="Autor",
            reach=10,
            url="http://example.com/articulo",
        )

        resultado = self.view._serializar_articulo(articulo, registro, None)

        self.assertEqual(resultado["fecha"], "2025-09-23 08:00:38 AM")

    def test_serializar_red_usa_formato_legible(self):
        registro = {
            "proveedor": "redes_twk",
            "datos_adicionales": {},
            "fecha": "2025-09-23T16:01:38.000Z",
            "red_social": "twitter.com",
        }
        red = SimpleNamespace(
            id="red-id",
            contenido="Contenido",
            fecha_publicacion=None,
            autor="Autor",
            reach=5,
            engagement=2,
            url="http://example.com/red",
            red_social=None,
        )

        resultado = self.view._serializar_red(red, registro, "red")

        self.assertEqual(resultado["fecha"], "2025-09-23 4:01:38 PM")
