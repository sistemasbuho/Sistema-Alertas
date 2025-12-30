import csv
import io
import logging
import os
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlparse

import requests
from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.response import Response
from rest_framework.views import APIView
from django.urls import NoReverseMatch, reverse

from apps.base.models import Articulo, DetalleEnvio, Redes, RedesSociales
from apps.proyectos.models import Proyecto
from apps.whatsapp.api.enviar_mensaje import enviar_alertas_automatico

from .contenido_redes import ajustar_contenido_red_social
from .utils import (
    combinar_fecha_hora,
    filtrar_registros_por_palabras,
    formatear_fecha_respuesta,
    limpiar_texto,
    normalizar_url,
    normalizar_valor_adicional,
    parsear_datetime,
    parsear_entero,
)


logger = logging.getLogger(__name__)


COLUMNAS_MEDIOS_TWK = {
    "title",
    "content",
    "published",
    "extra_author_attributes.name",
    "reach",
}

COLUMNAS_MEDIOS_TWK_VARIANTE_FUENTE = (
    COLUMNAS_MEDIOS_TWK - {"extra_author_attributes.name"}
) | {"extra_source_attributes.name"}

COLUMNAS_MEDIOS_GLOBAL_NEWS = {
    "autor - conductor",
    "medio",
    "fecha",
    "resumen - aclaracion",
    "resumen - aclaración",
    "título",
    "titulo",
    "audiencia",
}

COLUMNAS_MEDIOS_STAKEHOLDERS = {
    "autor",
    "fuente",
    "fecha",
    "resumen",
    "titular",
    "titulo",
    "audiencia",
}

COLUMNAS_MEDIOS_DETERM = {
    "author",
    "from",
    "title",
    "mention_snippet",
    "date",
}

COLUMNAS_REDES_TWK = {
    "content",
    "published",
    "extra_author_attributes.name",
    "reach",
    "engagement",
}

COLUMNAS_REDES_TWK_VARIANTE_AUTOR = (
    COLUMNAS_REDES_TWK - {"extra_author_attributes.name"}
) | {"extra_author_attributes.short_name"}

COLUMNAS_DETERM = {
    "mention_snippet",
    "date",
    "time",
    "reach",
    "engagement_rate",
    "author",
}

PROVEEDORES_NOMBRES = {
    "medios": "medios_twk",
    "redes": "redes_twk",
    "determ": "determ",
    "global_news": "global_news",
    "stakeholders": "stakeholders",
    "determ_medios": "determ_medios",
}

PROVEEDORES_ENDPOINTS = {
    "medios": "medios-alertas-ingestion",
    "redes": "redes-alertas-ingestion",
    "determ": "redes-alertas-ingestion",
    "global_news": "medios-alertas-ingestion",
    "stakeholders": "medios-alertas-ingestion",
    "determ_medios": "medios-alertas-ingestion",
}

DOMINIOS_REDES_SOCIALES = {
    "facebook.com": "Facebook",
    "twitter.com": "Twitter",
    "x.com": "Twitter",
    "instagram.com": "Instagram",
    "tiktok.com": "TikTok",
    "youtube.com": "YouTube",
    "linkedin.com": "LinkedIn",
}

CAMPOS_PRINCIPALES = {
    "medios": (
        COLUMNAS_MEDIOS_TWK
        | COLUMNAS_MEDIOS_GLOBAL_NEWS
        | COLUMNAS_MEDIOS_STAKEHOLDERS
        | COLUMNAS_MEDIOS_DETERM
        | {"url", "link", "extra_source_attributes.name"}
    ),
    "redes": (
        COLUMNAS_REDES_TWK
        | COLUMNAS_REDES_TWK_VARIANTE_AUTOR
        | {"url", "link", "red_social"}
    ),
    "determ": COLUMNAS_DETERM | {"url", "social_network"},
}


class IngestionAPIView(APIView):
    # authentication_classes: list = []
    # permission_classes: list = []

    def post(self, request):
        proyecto = self._obtener_proyecto(request)
        if proyecto is None:
            return Response({"detail": "Proyecto no encontrado o no indicado."}, status=400)

        tipo_alerta_proyecto = self._obtener_tipo_alerta_proyecto(proyecto)
        registros_estandar, proveedor, error_response = self._extraer_registros_estandar(
            request,
            tipo_alerta_proyecto,
        )
        if error_response:
            return error_response

        registros_filtrados = self._filtrar_por_criterios(registros_estandar, proyecto)

        if not registros_filtrados:
            respuesta = self._construir_respuesta_sin_registros(
                proveedor,
                proyecto,
            )
            self._notificar_ruta_externa(respuesta)
            return Response(respuesta, status=405)

        self._usuario_sistema_cache = self._obtener_usuario_desde_request(request)
        resultado = self._persistir_registros(registros_filtrados, proyecto)
        respuesta = self._construir_respuesta_exito(
            registros_filtrados,
            resultado,
            proveedor,
            proyecto,
        )

        self._procesar_envio_automatico(proyecto, respuesta)
        self._notificar_ruta_externa(respuesta)

        return Response(
            respuesta,
            status=201 if resultado["listado"] else 400,
        )

    def _extraer_registros_estandar(
        self,
        request,
        tipo_alerta_proyecto: Optional[str],
    ) -> Tuple[List[Dict[str, Any]], Optional[str], Optional[Response]]:
        registro_manual = self._obtener_registro_manual(request)
        if registro_manual:
            self._ajustar_registro_manual_por_tipo_alerta(
                registro_manual, tipo_alerta_proyecto
            )
            return [registro_manual], registro_manual.get("proveedor"), None

        archivos = self._obtener_archivos(request)
        if not archivos:
            return [], None, Response(
                {"detail": "Se requiere un archivo CSV o XLSX."},
                status=400,
            )

        registros_acumulados: List[Dict[str, Any]] = []
        proveedores_detectados: List[str] = []

        for archivo in archivos:
            extension = os.path.splitext(archivo.name)[1].lower()
            if extension not in {".csv", ".xlsx"}:
                return [], None, Response(
                    {"detail": "Formato de archivo no soportado."},
                    status=400,
                )

            headers, rows = self._parse_file(archivo, extension)
            headers, rows = self._normalizar_columnas_url(headers, rows)
            if not headers:
                return [], None, Response(
                    {"detail": "El archivo no contiene encabezados válidos."},
                    status=400,
                )

            error_url = self._validar_columna_url(headers, rows)
            if error_url:
                return [], None, error_url

            provider = self._detectar_proveedor(headers)
            if not provider:
                return [], None, Response(
                    {
                        "detail": "No fue posible determinar el tipo de datos del archivo.",
                    },
                    status=400,
                )

            # Validar que el proveedor detectado coincida con el tipo de alerta del proyecto
            error_validacion = self._validar_tipo_archivo_con_proyecto(provider, tipo_alerta_proyecto)
            if error_validacion:
                return [], provider, error_validacion

            registros_estandar = self._mapear_filas(provider, rows)
            if not registros_estandar:
                return [], provider, Response(
                    {"detail": "No se encontraron filas válidas en el archivo."},
                    status=400,
                )

            registros_acumulados.extend(registros_estandar)
            proveedores_detectados.append(provider)

        provider_final: Optional[str] = None
        if proveedores_detectados:
            provider_final = proveedores_detectados[0]
            for proveedor in proveedores_detectados[1:]:
                if proveedor != provider_final:
                    provider_final = "multiple"
                    break

        return registros_acumulados, provider_final, None

    def _construir_respuesta_sin_registros(
        self,
        proveedor: Optional[str],
        proyecto: Proyecto,
    ) -> Dict[str, Any]:
        return {
            "proveedor": proveedor,
            "mensaje": "0 registros cumplen con los criterios de aceptación configurados.",
            "listado": [],
            "errores": [],
            "duplicados": 0,
            "descartados": 0,
            "proyecto_keywords": self._obtener_keywords_proyecto(proyecto),
            "proyecto_nombre": self._obtener_nombre_proyecto(proyecto),
        }

    def _construir_respuesta_exito(
        self,
        registros_filtrados: List[Dict[str, Any]],
        resultado: Dict[str, List[Dict[str, Any]]],
        proveedor: Optional[str],
        proyecto: Proyecto,
    ) -> Dict[str, Any]:
        proveedor_respuesta = registros_filtrados[0].get("proveedor") or proveedor
        total_creados = len(resultado["listado"])
        duplicados = resultado.get("duplicados", 0)
        descartados = resultado.get("descartados", 0)
        detalles: List[str] = []
        if duplicados:
            detalles.append(f"{duplicados} duplicados")
        if descartados:
            detalles.append(f"{descartados} descartados")

        mensaje = f"{total_creados} registros creados"
        if detalles:
            mensaje = f"{mensaje} ({', '.join(detalles)})"

        return {
            "proveedor": proveedor_respuesta,
            "mensaje": mensaje,
            "listado": resultado["listado"],
            "errores": resultado["errores"],
            "duplicados": duplicados,
            "descartados": descartados,
            "proyecto_keywords": self._obtener_keywords_proyecto(proyecto),
            "proyecto_nombre": self._obtener_nombre_proyecto(proyecto),
        }

    # ------------------------------------------------------------------
    # Extracción de datos de la request
    # ------------------------------------------------------------------
    def _obtener_proyecto(self, request) -> Optional[Proyecto]:
        posibles_ids: Iterable[str] = (
            request.data.get("proyecto_id"),
            request.data.get("proyecto"),
            request.POST.get("proyecto_id"),
            request.POST.get("proyecto"),
            request.query_params.get("proyecto"),
        )
        proyecto_id = next((pid for pid in posibles_ids if pid), None)
        if not proyecto_id:
            return None
        return Proyecto.objects.filter(id=proyecto_id).first()

    def _obtener_archivos(self, request) -> List[Any]:
        files = getattr(request, "FILES", None)
        if not files:
            return []

        archivos: List[Any] = []

        def _agregar(valor: Any) -> None:
            if not valor:
                return
            if isinstance(valor, (list, tuple, set)):
                for item in valor:
                    _agregar(item)
                return
            archivos.append(valor)

        if hasattr(files, "lists"):
            for _, valores in files.lists():  # type: ignore[attr-defined]
                _agregar(valores)
        elif hasattr(files, "values"):
            for valor in files.values():  # type: ignore[attr-defined]
                _agregar(valor)
        else:
            _agregar(files)

        if not archivos and hasattr(files, "get"):
            for clave in ("file", "archivo", "archivos", "files"):
                valor = files.get(clave)  # type: ignore[attr-defined]
                if valor:
                    _agregar(valor)

        archivos_unicos: List[Any] = []
        vistos = set()
        for archivo in archivos:
            identificador = id(archivo)
            if identificador in vistos:
                continue
            vistos.add(identificador)
            archivos_unicos.append(archivo)
        return archivos_unicos

    def _obtener_registro_manual(self, request) -> Optional[Dict[str, Any]]:
        data_sources = [request.data, request.POST]
        for data in data_sources:
            if not hasattr(data, "get"):
                continue

            url = normalizar_url(self._obtener_valor_data(data, "url") or self._obtener_valor_data(data, "link"))
            if not url:
                continue

            tipo = (self._obtener_valor_data(data, "tipo") or "articulo").strip().lower()
            if tipo not in {"articulo", "red"}:
                tipo = "articulo"

            fecha_raw = (
                self._obtener_valor_data(data, "fecha")
                or self._obtener_valor_data(data, "published")
                or self._obtener_valor_data(data, "fecha_publicacion")
            )

            red_social_valor = limpiar_texto(
                self._obtener_valor_data(data, "red_social")
                or self._obtener_valor_data(data, "social_network")
            )

            if not red_social_valor and url:
                parsed_url = urlparse(url)
                domain = parsed_url.netloc.lower()
                for dominio, nombre in DOMINIOS_REDES_SOCIALES.items():
                    if dominio in domain:
                        red_social_valor = domain
                        break

            registro: Dict[str, Any] = {
                "tipo": tipo,
                "titulo": limpiar_texto(
                    self._obtener_valor_data(data, "titulo")
                    or self._obtener_valor_data(data, "title")
                ),
                "contenido": limpiar_texto(
                    self._obtener_valor_data(data, "contenido")
                    or self._obtener_valor_data(data, "content")
                ),
                "fecha": parsear_datetime(fecha_raw) if fecha_raw else None,
                "autor": limpiar_texto(
                    self._obtener_valor_data(data, "autor")
                    or self._obtener_valor_data(data, "extra_source_attributes.name")
                    or self._obtener_valor_data(data, "extra_author_attributes.short_name")
                    or self._obtener_valor_data(data, "extra_author_attributes.name")
                ),
                "reach": parsear_entero(self._obtener_valor_data(data, "reach")),
                "engagement": parsear_entero(self._obtener_valor_data(data, "engagement")),
                "url": url,
                "red_social": red_social_valor,
                "proveedor": "manual",
                "datos_adicionales": {},
            }

            if registro["tipo"] == "red":
                registro["contenido"] = ajustar_contenido_red_social(
                    registro.get("contenido"), registro.get("red_social")
                )

            adicionales = {}
            for clave, valor in self._iterar_items_data(data):
                if clave in {
                    "url",
                    "link",
                    "proyecto",
                    "proyecto_id",
                    "tipo",
                    "titulo",
                    "title",
                    "contenido",
                    "content",
                    "fecha",
                    "published",
                    "fecha_publicacion",
                    "autor",
                    "extra_source_attributes.name",
                    "extra_author_attributes.name",
                    "extra_author_attributes.short_name",
                    "reach",
                    "engagement",
                    "red_social",
                    "social_network",
                }:
                    continue
                valor_normalizado = normalizar_valor_adicional(valor)
                if valor_normalizado is not None:
                    adicionales[clave] = valor_normalizado

            if adicionales:
                registro["datos_adicionales"] = adicionales

            return registro

        return None

    def _ajustar_registro_manual_por_tipo_alerta(
        self,
        registro: Dict[str, Any],
        tipo_alerta: Optional[str],
    ) -> None:
        if not tipo_alerta:
            return

        tipo_alerta_normalizado = tipo_alerta.strip().lower()
        if not tipo_alerta_normalizado:
            return

        if tipo_alerta_normalizado == "redes":
            registro["tipo"] = "red"
        elif tipo_alerta_normalizado == "medios":
            registro["tipo"] = "articulo"

    def _obtener_valor_data(self, data, key):
        valor = data.get(key)  # type: ignore[attr-defined]
        if isinstance(valor, list):
            return valor[0]
        return valor

    def _iterar_items_data(self, data):
        if hasattr(data, "lists"):
            for clave, valores in data.lists():  # type: ignore[attr-defined]
                if not valores:
                    continue
                yield clave, valores[0] if len(valores) == 1 else valores
        elif hasattr(data, "items"):
            yield from data.items()  # type: ignore[attr-defined]

    # ------------------------------------------------------------------
    # Procesamiento de archivos
    # ------------------------------------------------------------------
    def _parse_file(self, uploaded_file, extension: str) -> Tuple[List[str], List[Dict[str, Any]]]:
        if extension == ".csv":
            return self._parse_csv(uploaded_file)
        return self._parse_xlsx(uploaded_file)

    def _parse_csv(self, uploaded_file) -> Tuple[List[str], List[Dict[str, Any]]]:
        uploaded_file.seek(0)
        data = uploaded_file.read().decode("utf-8-sig")
        csv_buffer = io.StringIO(data)
        reader = csv.DictReader(csv_buffer)
        headers = [self._normalizar_encabezado(h) for h in (reader.fieldnames or [])]
        rows = []
        for raw_row in reader:
            rows.append({self._normalizar_encabezado(k): v for k, v in raw_row.items()})
        return headers, rows

    def _parse_xlsx(self, uploaded_file) -> Tuple[List[str], List[Dict[str, Any]]]:
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=False)
        headers_row = None
        for potential_header_row in rows_iter:
            if any(
                self._valor_contiene_datos(cell.value) if cell is not None else False
                for cell in potential_header_row
            ):
                headers_row = potential_header_row
                break

        if headers_row is None:
            return [], []

        headers = [self._normalizar_encabezado(cell.value) for cell in headers_row]
        rows: List[Dict[str, Any]] = []
        header_indices = [
            (index, header)
            for index, header in enumerate(headers)
            if header
        ]
        columnas_con_datos = {header: False for _, header in header_indices}

        for row in rows_iter:
            row_dict: Dict[str, Any] = {}
            for idx, header in header_indices:
                cell = row[idx] if idx < len(row) else None
                value = None
                if cell is not None:
                    value = cell.value
                    hyperlink = getattr(cell, "hyperlink", None)
                    if hyperlink:
                        value = hyperlink.target or hyperlink.location or value
                if self._valor_contiene_datos(value):
                    columnas_con_datos[header] = True
                    row_dict[header] = value
            if row_dict:
                rows.append(row_dict)

        columnas_vacias = {
            header for header, tiene_datos in columnas_con_datos.items() if not tiene_datos
        }
        if columnas_vacias:
            for row_dict in rows:
                for columna in columnas_vacias:
                    row_dict.pop(columna, None)
        return headers, rows

    def _normalizar_columnas_url(
        self, headers: List[str], rows: List[Dict[str, Any]]
    ) -> Tuple[List[str], List[Dict[str, Any]]]:
        headers_normalizados = [header for header in headers if header]
        if "url" in headers_normalizados:
            return headers, rows

        columnas_alternativas = [
            "link (streaming - imagen)",
            "link (streaming – imagen)",
            "link",
        ]

        for columna in columnas_alternativas:
            if columna not in headers_normalizados:
                continue
            if "url" not in headers:
                headers.append("url")
            for row in rows:
                valor_url = row.get("url")
                if valor_url:
                    continue
                valor_alternativo = row.get(columna)
                if valor_alternativo:
                    row["url"] = valor_alternativo
            return headers, rows

        return headers, rows

    def _normalizar_encabezado(self, header_value) -> str:
        if header_value is None:
            return ""
        return str(header_value).strip().lower()

    def _valor_contiene_datos(self, value: Any) -> bool:
        if value is None:
            return False
        if isinstance(value, str):
            return bool(value.strip())
        return True

    # ------------------------------------------------------------------
    # Detección y mapeo de filas
    # ------------------------------------------------------------------
    def _detectar_proveedor(self, headers: List[str]) -> str:
        header_set = set(headers)
        if header_set >= COLUMNAS_MEDIOS_TWK or header_set >= COLUMNAS_MEDIOS_TWK_VARIANTE_FUENTE:
            return "medios"
        if header_set >= COLUMNAS_REDES_TWK or header_set >= COLUMNAS_REDES_TWK_VARIANTE_AUTOR:
            return "redes"
        if self._headers_corresponden_a_determ_medios(header_set):
            return "determ_medios"
        if header_set >= COLUMNAS_DETERM:
            return "determ"
        if self._headers_corresponden_a_global_news(header_set):
            return "global_news"
        if self._headers_corresponden_a_stakeholders(header_set):
            return "stakeholders"
        return ""

    def _validar_columna_url(
        self, headers: List[str], rows: List[Dict[str, Any]]
    ) -> Optional[Response]:
        headers_normalizados = {header for header in headers if header}
        if "url" not in headers_normalizados:
            return Response(
                {"detail": "El archivo debe incluir una columna 'url'."},
                status=400,
            )

        for row in rows:
            url_valida = normalizar_url(row.get("url")) if isinstance(row, dict) else None
            if url_valida:
                return None

        return Response(
            {"detail": "La columna 'url' debe contener al menos un valor válido."},
            status=400,
        )

    def _validar_tipo_archivo_con_proyecto(
        self, provider: str, tipo_alerta_proyecto: Optional[str]
    ) -> Optional[Response]:
        # Validación deshabilitada - se valida por tipo de URL en _inferir_proveedor
        return None

    def _headers_corresponden_a_global_news(self, header_set: set) -> bool:
        requeridos = {"autor - conductor", "medio", "fecha"}
        if not requeridos.issubset(header_set):
            return False
        if not {"resumen - aclaracion", "resumen - aclaración", "resumen"} & header_set:
            return False
        if not {"título", "titulo", "titular"} & header_set:
            return False
        return True

    def _headers_corresponden_a_stakeholders(self, header_set: set) -> bool:
        requeridos = {"autor", "fuente", "fecha", "resumen"}
        if not requeridos.issubset(header_set):
            return False
        if not {"titular", "titulo", "título"} & header_set:
            return False
        return True

    def _headers_corresponden_a_determ_medios(self, header_set: set) -> bool:
        requeridos = {"author", "from", "title", "mention_snippet", "date"}
        return requeridos.issubset(header_set)

    def _obtener_nombre_proveedor(self, provider: str) -> str:
        return PROVEEDORES_NOMBRES.get(provider, provider)

    def _mapear_filas(self, provider: str, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        if provider == "determ":
            map_function = self._mapear_determ
            campos_principales = CAMPOS_PRINCIPALES.get("determ", set())
            nombre_proveedor = self._obtener_nombre_proveedor("determ")
            registros: List[Dict[str, Any]] = []
            for row in rows:
                registro = map_function(row)
                registro["proveedor"] = nombre_proveedor
                registro["datos_adicionales"] = self._extraer_datos_adicionales(row, campos_principales)
                registros.append(registro)
            return registros

        registros = []
        for row in rows:
            proveedor_inferido = self._inferir_proveedor(row)
            campos_principales = CAMPOS_PRINCIPALES.get(proveedor_inferido, set())
            proveedor_respuesta = (
                provider if provider in PROVEEDORES_NOMBRES else proveedor_inferido
            )
            nombre_proveedor = self._obtener_nombre_proveedor(proveedor_respuesta)
            if proveedor_inferido == "redes":
                registro = self._mapear_redes_twk(row)
            else:
                registro = self._mapear_medios_twk(row, proveedor_respuesta)
            registro["proveedor"] = nombre_proveedor
            registro["datos_adicionales"] = self._extraer_datos_adicionales(row, campos_principales)
            registros.append(registro)
        return registros

    def _inferir_proveedor(self, row: Dict[str, Any]) -> str:
        if row.get("red_social"):
            return "redes"

        domain_url = row.get("domain_url")
        if domain_url:
            domain_normalized = str(domain_url).lower()
            if any(
                dominio in domain_normalized
                for dominio in DOMINIOS_REDES_SOCIALES.keys()
            ):
                return "redes"

        return "medios"

    def _filtrar_por_criterios(
        self, registros: List[Dict[str, Any]], proyecto: Proyecto
    ) -> List[Dict[str, Any]]:
        criterios: Iterable[str] = []
        if proyecto and hasattr(proyecto, "get_criterios_aceptacion_list"):
            criterios = proyecto.get_criterios_aceptacion_list()
        return filtrar_registros_por_palabras(registros, criterios)

    def _obtener_keywords_proyecto(self, proyecto: Optional[Proyecto]) -> List[str]:
        if not proyecto:
            return []

        if hasattr(proyecto, "get_keywords_list"):
            keywords = proyecto.get_keywords_list()
            if isinstance(keywords, list):
                return keywords

        keywords_attr = getattr(proyecto, "keywords", None)
        if keywords_attr:
            return [
                keyword.strip()
                for keyword in str(keywords_attr).split(",")
                if keyword and keyword.strip()
            ]

        return []

    def _obtener_nombre_proyecto(self, proyecto: Optional[Proyecto]) -> str:
        if not proyecto:
            return ""

        nombre = getattr(proyecto, "nombre", "")
        if not nombre:
            return ""

        return str(nombre)

    def _construir_payload_forward(
        self,
        provider: str,
        registros: List[Dict[str, Any]],
        proyecto: Proyecto,
    ) -> Dict[str, Any]:
        proveedor_nombre = registros[0].get("proveedor") or self._obtener_nombre_proveedor(provider)
        tipo_alerta_proyecto = self._obtener_tipo_alerta_proyecto(proyecto)
        alertas = []
        for registro in registros:
            fecha_legible = formatear_fecha_respuesta(registro.get("fecha"))
            alerta = {
                "tipo": tipo_alerta_proyecto or registro.get("tipo"),
                "titulo": registro.get("titulo"),
                "contenido": registro.get("contenido"),
                "fecha": fecha_legible,
                "autor": registro.get("autor"),
                "reach": registro.get("reach"),
                "engagement": registro.get("engagement"),
                "url": registro.get("url"),
                "red_social": registro.get("red_social"),
                "datos_adicionales": registro.get("datos_adicionales") or {},
            }
            for campo in ("reach", "engagement"):
                valor = alerta.get(campo)
                if valor is not None and not isinstance(valor, str):
                    alerta[campo] = str(valor)
            alertas.append(alerta)

        return {
            "proveedor": proveedor_nombre,
            "proyecto": str(proyecto.id),
            "alertas": alertas,
        }

    def _mapear_medios_twk(
        self, row: Dict[str, Any], provider: Optional[str] = None
    ) -> Dict[str, Any]:
        provider_normalized = (provider or "").strip().lower()
        fecha = None
        fecha_raw: Optional[Any] = None
        hora_raw: Optional[Any] = None

        if provider_normalized == "global_news":
            hora_candidata = row.get("Hora") or row.get("hora")
            if self._valor_contiene_datos(hora_candidata):
                hora_raw = hora_candidata

        if provider_normalized == "medios":
            fecha = parsear_datetime(row.get("published"))
        elif provider_normalized == "global_news":
            fecha_raw = self._obtener_primera_coincidencia(row, ["fecha"])
            if hora_raw is not None:
                fecha = combinar_fecha_hora(fecha_raw, hora_raw)
            if fecha is None:
                fecha = parsear_datetime(fecha_raw)

        if fecha is None:
            if fecha_raw is None:
                fecha_raw = self._obtener_primera_coincidencia(
                    row,
                    ["fecha", "published", "date"],
                )

            if hora_raw is None and provider_normalized != "stakeholders":
                hora_raw = self._obtener_primera_coincidencia(row, ["hora", "time"])

            if hora_raw is not None:
                fecha = combinar_fecha_hora(fecha_raw, hora_raw)

            if fecha is None:
                fecha = parsear_datetime(fecha_raw)
        titulo = limpiar_texto(
            self._obtener_primera_coincidencia(
                row,
                ["title", "título", "titulo", "titular"],
            )
        )
        contenido_valor: Optional[Any]
        if provider_normalized == "medios":
            contenido_valor = row.get("content_snippet")
        elif provider_normalized == "global_news":
            contenido_valor = self._obtener_primera_coincidencia(
                row,
                ["resumen - aclaracion", "resumen - aclaración"],
            )
        elif provider_normalized == "stakeholders":
            contenido_valor = row.get("resumen")
        else:
            contenido_valor = None

        if not self._valor_contiene_datos(contenido_valor):
            contenido_valor = self._obtener_primera_coincidencia(
                row,
                [
                    "content_snippet",
                    "content",
                    "resumen",
                    "resumen - aclaracion",
                    "resumen - aclaración",
                    "mention_snippet",
                ],
            )

        contenido = limpiar_texto(contenido_valor)

        # Obtener fuente según proveedor
        fuente_valor: Optional[Any] = None
        if provider_normalized == "global_news":
            fuente_valor = self._obtener_primera_coincidencia(row, ["Medio", "medio"])
        elif provider_normalized == "stakeholders":
            fuente_valor = self._obtener_primera_coincidencia(row, ["Fuente", "fuente"])

        fuente = limpiar_texto(fuente_valor) if fuente_valor else None

        # Obtener tipo_medio según proveedor
        tipo_medio: Optional[str] = None
        if provider_normalized == "global_news":
            tipo_medio_raw = self._obtener_primera_coincidencia(row, ["Tipo de Medio", "tipo de medio"])
            if tipo_medio_raw:
                tipo_medio_str = str(tipo_medio_raw).strip().lower()
                if "cable" in tipo_medio_str:
                    tipo_medio = "Televisión"
                elif "fm" in tipo_medio_str:
                    tipo_medio = "Radio"
                elif "diario" in tipo_medio_str or "revista" in tipo_medio_str:
                    tipo_medio = "Prensa"
                else:
                    tipo_medio = limpiar_texto(tipo_medio_raw)
        elif provider_normalized == "stakeholders":
            medio_raw = self._obtener_primera_coincidencia(row, ["Medio", "medio"])
            if medio_raw:
                medio_str = str(medio_raw).strip().lower()
                if "internet" in medio_str:
                    tipo_medio = "Online"
                else:
                    tipo_medio = limpiar_texto(medio_raw)
        elif provider_normalized in {"medios", "determ_medios"}:
            tipo_medio = "Online"

        if provider_normalized == "medios":
            autor_valor = self._obtener_primera_coincidencia(
                row,
                [
                    "extra_source_attributes.name",
                    "extra_author_attributes.short_name",
                    "extra_author_attributes.name",
                    "autor - conductor",
                    "autor",
                    "author",
                ],
            )
        elif provider_normalized == "global_news":
            autor_valor = self._obtener_primera_coincidencia(
                row,
                [
                    "Medio",
                    "medio",
                    "autor - conductor",
                    "autor",
                    "author",
                ],
            )
        elif provider_normalized == "stakeholders":
            autor_valor = self._obtener_primera_coincidencia(
                row,
                [
                    "Fuente",
                    "fuente",
                    "autor",
                    "autor - conductor",
                    "author",
                ],
            )
        elif provider_normalized == "determ_medios":
            autor_valor = row.get("from")
            if not self._valor_contiene_datos(autor_valor):
                autor_valor = row.get("FROM")
        else:
            autor_valor = self._obtener_primera_coincidencia(
                row,
                [
                    "extra_source_attributes.name",
                    "extra_author_attributes.short_name",
                    "extra_author_attributes.name",
                    "autor - conductor",
                    "autor",
                    "author",
                ],
            )

        autor = limpiar_texto(autor_valor)
        reach_claves = ["reach"]
        if provider_normalized in {"global_news", "stakeholders"}:
            reach_claves.insert(0, "audiencia")

        reach = parsear_entero(
            self._obtener_primera_coincidencia(row, reach_claves)
        )
        url = normalizar_url(
            self._obtener_primera_coincidencia(
                row,
                [
                    "url",
                    "link",
                    "link (streaming - imagen)",
                    "link (streaming – imagen)",
                ],
            )
        )
        engagement = parsear_entero(
            self._obtener_primera_coincidencia(row, ["engagement", "engagement_rate"])
        )
        return {
            "tipo": "articulo",
            "titulo": titulo,
            "contenido": contenido,
            "fecha": fecha,
            "autor": autor,
            "fuente": fuente,
            "tipo_medio": tipo_medio,
            "reach": reach,
            "engagement": engagement,
            "url": url,
        }

    def _mapear_redes_twk(self, row: Dict[str, Any]) -> Dict[str, Any]:
        fecha = parsear_datetime(row.get("published"))
        red_social = limpiar_texto(row.get("domain_url"))
        contenido = ajustar_contenido_red_social(
            limpiar_texto(row.get("content")),
            red_social,
        )
        return {
            "tipo": "red",
            "contenido": contenido,
            "fecha": fecha,
            "autor": limpiar_texto(
                row.get("extra_author_attributes.short_name")
                or row.get("extra_author_attributes.name")
            ),
            "reach": parsear_entero(row.get("reach")),
            "engagement": parsear_entero(row.get("engagement")),
            "url": normalizar_url(row.get("url") or row.get("link")),
            "red_social": red_social,
        }

    def _mapear_determ(self, row: Dict[str, Any]) -> Dict[str, Any]:
        fecha = combinar_fecha_hora(row.get("date"), row.get("time"))
        red_social = limpiar_texto(row.get("social_network"))
        contenido = ajustar_contenido_red_social(
            limpiar_texto(row.get("mention_snippet")),
            red_social,
        )
        return {
            "tipo": "red",
            "contenido": contenido,
            "fecha": fecha,
            "autor": limpiar_texto(
                row.get("author")
                or row.get("AUTHOR")
                or row.get("FROM")
                or row.get("from")
            ),
            "reach": parsear_entero(row.get("reach")),
            "engagement": parsear_entero(row.get("engagement_rate")),
            "url": normalizar_url(row.get("url")),
            "red_social": red_social,
        }

    def _obtener_primera_coincidencia(
        self, row: Dict[str, Any], claves: Iterable[str]
    ) -> Optional[Any]:
        for clave in claves:
            if clave not in row:
                continue
            valor = row.get(clave)
            if valor in (None, ""):
                continue
            return valor
        return None

    # ------------------------------------------------------------------
    # Persistencia y serialización
    # ------------------------------------------------------------------
    def _obtener_usuario_desde_request(self, request):
        usuario = getattr(request, "user", None)
        if usuario and getattr(usuario, "is_authenticated", False):
            return usuario

        posibles_fuentes = []
        if hasattr(request, "data"):
            posibles_fuentes.append(request.data)
        if hasattr(request, "query_params"):
            posibles_fuentes.append(request.query_params)

        UserModel = get_user_model()

        for fuente in posibles_fuentes:
            if not hasattr(fuente, "get"):
                continue
            for clave in ("usuario_id", "usuario", "user_id", "created_by"):
                valor = fuente.get(clave)
                if isinstance(valor, list):
                    valor = valor[0]
                if not valor:
                    continue
                try:
                    return UserModel.objects.get(id=valor)
                except (UserModel.DoesNotExist, ValueError, TypeError):
                    continue

        return None

    def _persistir_registros(self, registros: List[Dict[str, Any]], proyecto: Proyecto) -> Dict[str, List[Dict[str, Any]]]:
        errores: List[Dict[str, Any]] = []
        listado: List[Dict[str, Any]] = []
        duplicados = 0
        descartados = 0
        sistema_user = getattr(self, "_usuario_sistema_cache", None)
        tipo_alerta_proyecto = self._obtener_tipo_alerta_proyecto(proyecto)

        for indice, registro in enumerate(registros, start=1):
            try:
                if registro.get("tipo") == "articulo":
                    articulo = self._crear_articulo(registro, proyecto, sistema_user)
                    listado.append(
                        self._serializar_articulo(articulo, registro, tipo_alerta_proyecto)
                    )
                else:
                    red = self._crear_red_social(registro, proyecto)
                    listado.append(
                        self._serializar_red(red, registro, tipo_alerta_proyecto)
                    )
            except Exception as exc:  # pylint: disable=broad-except
                logger.exception("Error procesando fila %s", indice)
                mensaje_error = str(exc)
                if isinstance(exc, ValueError) and "ya existe" in mensaje_error.lower():
                    duplicados += 1
                else:
                    descartados += 1
                errores.append({"fila": indice, "error": mensaje_error})
        return {
            "listado": listado,
            "errores": errores,
            "duplicados": duplicados,
            "descartados": descartados,
        }

    def _asegurar_detalle_envio(
        self,
        *,
        proyecto: Proyecto,
        usuario,
        articulo: Optional[Articulo] = None,
        red: Optional[Redes] = None,
    ) -> DetalleEnvio:
        if (articulo is None and red is None) or (articulo is not None and red is not None):
            raise ValueError("Debe especificar únicamente un medio o una red social")

        filtros: Dict[str, Any] = {"proyecto": proyecto}
        if articulo is not None:
            filtros["medio"] = articulo
        else:
            filtros["red_social"] = red

        defaults: Dict[str, Any] = {
            "estado_enviado": False,
            "estado_revisado": True,
        }

        if usuario:
            defaults["created_by"] = usuario
            defaults["modified_by"] = usuario

        detalle, creado = DetalleEnvio.objects.get_or_create(
            **filtros,
            defaults=defaults,
        )

        if not creado:
            actualizaciones: Dict[str, Any] = {"estado_revisado": True}
            if usuario:
                actualizaciones["modified_by"] = usuario
            DetalleEnvio.objects.filter(pk=detalle.pk).update(**actualizaciones)
            detalle.refresh_from_db()

        return detalle

    def _crear_articulo(self, registro: Dict[str, Any], proyecto: Proyecto, sistema_user) -> Articulo:
        with transaction.atomic():
            url = registro.get("url") or ""
            if self._es_url_duplicada_por_proyecto(Articulo, proyecto, url):
                raise ValueError("La URL ya existe para este proyecto")

            kwargs: Dict[str, Any] = {
                "titulo": registro.get("titulo"),
                "contenido": registro.get("contenido"),
                "url": url,
                "fecha_publicacion": registro.get("fecha") or timezone.now(),
                "autor": registro.get("autor"),
                "fuente": registro.get("fuente"),
                "tipo_medio": registro.get("tipo_medio"),
                "reach": registro.get("reach"),
                "engagement": registro.get("engagement"),
                "proyecto": proyecto,
            }

            if sistema_user:
                kwargs["created_by"] = sistema_user
                kwargs["modified_by"] = sistema_user

            articulo = Articulo.objects.create(**kwargs)

            self._asegurar_detalle_envio(
                articulo=articulo,
                proyecto=proyecto,
                usuario=sistema_user,
            )
        return articulo

    def _crear_red_social(self, registro: Dict[str, Any], proyecto: Proyecto) -> Redes:
        with transaction.atomic():
            url = registro.get("url") or ""
            if self._es_url_duplicada_por_proyecto(Redes, proyecto, url):
                raise ValueError("La URL ya existe para este proyecto")

            red_social_obj = None
            nombre_red = registro.get("red_social")
            if nombre_red:
                domain_parse = urlparse(nombre_red)
                domain_candidate = (domain_parse.netloc or domain_parse.path or "").lower()
                if domain_candidate.startswith("www."):
                    domain_candidate = domain_candidate[4:]

                candidate_names: List[str] = []
                if domain_candidate:
                    for dominio, nombre in DOMINIOS_REDES_SOCIALES.items():
                        dominio_normalizado = dominio.lower()
                        if dominio_normalizado.startswith("www."):
                            dominio_normalizado = dominio_normalizado[4:]
                        dominio_base = dominio_normalizado.split(".")[0]
                        if dominio_normalizado and dominio_normalizado in domain_candidate:
                            candidate_names.append(nombre)
                        elif dominio_base and dominio_base in domain_candidate:
                            candidate_names.append(nombre)

                candidate_names.extend(
                    [
                        domain_candidate,
                        domain_candidate.split(".")[0] if domain_candidate else "",
                        str(nombre_red).strip(),
                    ]
                )

                for nombre_candidato in dict.fromkeys(
                    valor for valor in candidate_names if valor
                ):
                    red_social_obj = RedesSociales.objects.filter(
                        nombre__iexact=nombre_candidato
                    ).first()
                    if red_social_obj:
                        break

            usuario_creador = getattr(self, "_usuario_sistema_cache", None)

            kwargs: Dict[str, Any] = {
                "contenido": registro.get("contenido"),
                "fecha_publicacion": registro.get("fecha") or timezone.now(),
                "url": url,
                "autor": registro.get("autor"),
                "reach": registro.get("reach"),
                "engagement": registro.get("engagement"),
                "red_social": red_social_obj,
                "proyecto": proyecto,
            }

            if usuario_creador:
                kwargs["created_by"] = usuario_creador
                kwargs["modified_by"] = usuario_creador

            red = Redes.objects.create(**kwargs)

            self._asegurar_detalle_envio(
                red=red,
                proyecto=proyecto,
                usuario=usuario_creador,
            )
        return red

    def _procesar_envio_automatico(
        self,
        proyecto: Proyecto,
        respuesta: Dict[str, Any],
    ) -> None:
        if not proyecto:
            return

        tipo_envio = getattr(proyecto, "tipo_envio", "")
        if not isinstance(tipo_envio, str) or tipo_envio.strip().lower() != "automatico":
            return

        listado = respuesta.get("listado") or []
        if not listado:
            return

        tipo_alerta = self._obtener_tipo_alerta_proyecto(proyecto)
        if not tipo_alerta:
            tipo_alerta = next(
                (alerta.get("tipo") for alerta in listado if alerta.get("tipo")),
                None,
            )

        if not tipo_alerta:
            return

        alertas = [self._preparar_alerta_para_envio(alerta) for alerta in listado]

        usuario = getattr(self, "_usuario_sistema_cache", None)
        usuario_id = getattr(usuario, "id", None) or getattr(usuario, "pk", None)

        kwargs: Dict[str, Any] = {}
        if usuario_id:
            kwargs["usuario_id"] = usuario_id

        try:
            enviar_alertas_automatico(
                proyecto.id,
                tipo_alerta,
                alertas,
                **kwargs,
            )
        except Exception:  # pylint: disable=broad-except
            logger.exception(
                "Error enviando alertas automáticas para el proyecto %s",
                proyecto.id,
            )

    def _preparar_alerta_para_envio(self, alerta: Dict[str, Any]) -> Dict[str, Any]:
        alerta_envio = alerta.copy()
        datos_adicionales = alerta_envio.get("datos_adicionales")
        if isinstance(datos_adicionales, dict):
            alerta_envio["datos_adicionales"] = datos_adicionales.copy()
        return alerta_envio

    def _serializar_articulo(
        self,
        articulo: Articulo,
        registro: Dict[str, Any],
        tipo_alerta: Optional[str] = None,
    ) -> Dict[str, Any]:
        fecha_origen = articulo.fecha_publicacion or registro.get("fecha")
        return {
            "id": str(articulo.id),
            "tipo": (tipo_alerta or "articulo"),
            "titulo": articulo.titulo,
            "contenido": articulo.contenido,
            "fecha": formatear_fecha_respuesta(fecha_origen),
            "fecha_creacion": formatear_fecha_respuesta(articulo.created_at),
            "autor": articulo.autor,
            "reach": articulo.reach,
            "engagement": articulo.engagement,
            "url": articulo.url,
            "red_social": None,
            "proveedor": registro.get("proveedor"),
            "datos_adicionales": registro.get("datos_adicionales") or {},
        }

    def _serializar_red(
        self,
        red: Redes,
        registro: Dict[str, Any],
        tipo_alerta: Optional[str] = None,
    ) -> Dict[str, Any]:
        fecha_origen = red.fecha_publicacion or registro.get("fecha")
        return {
            "id": str(red.id),
            "tipo": (tipo_alerta or "red"),
            "titulo": None,
            "contenido": red.contenido,
            "fecha": formatear_fecha_respuesta(fecha_origen),
            "fecha_creacion": formatear_fecha_respuesta(red.created_at),
            "autor": red.autor,
            "reach": red.reach,
            "engagement": red.engagement,
            "url": red.url,
            "red_social": red.red_social.nombre if red.red_social else None,
            "proveedor": registro.get("proveedor"),
            "datos_adicionales": registro.get("datos_adicionales") or {},
        }

    # ------------------------------------------------------------------
    # Utilidades
    # ------------------------------------------------------------------
    def _obtener_tipo_alerta_proyecto(self, proyecto: Optional[Proyecto]) -> Optional[str]:
        if not proyecto:
            return None

        tipo_alerta = getattr(proyecto, "tipo_alerta", None)
        if isinstance(tipo_alerta, str):
            tipo_alerta_normalizado = tipo_alerta.strip()
            if tipo_alerta_normalizado:
                return tipo_alerta_normalizado
        return None

    def _obtener_usuario_sistema(self):
        UserModel = get_user_model()
        try:
            return UserModel.objects.get(id=2)
        except UserModel.DoesNotExist as exc:  # type: ignore[attr-defined]
            raise ValueError("El usuario del sistema (id=2) no existe") from exc

    def _extraer_datos_adicionales(self, row: Dict[str, Any], campos_principales: Iterable[str]) -> Dict[str, Any]:
        adicionales: Dict[str, Any] = {}
        for key, value in row.items():
            valor_limpio = normalizar_valor_adicional(value)
            if valor_limpio is not None:
                adicionales[key] = valor_limpio
        return adicionales

    def _notificar_ruta_externa(self, payload: Dict[str, Any]) -> None:
        url = getattr(settings, "RUTA_X_URL", None) or "http://localhost:8000/ruta_x"
        try:
            requests.post(url, json=payload, timeout=5)
        except requests.RequestException as exc:  # pylint: disable=broad-except
            logger.warning("No fue posible notificar la ruta externa %s: %s", url, exc)

    # ------------------------------------------------------------------
    # Validación de URLs por proyecto
    # ------------------------------------------------------------------
    def _es_url_duplicada_por_proyecto(self, model, proyecto: Proyecto, url: Optional[str]) -> bool:
        if not url:
            return False

        clave_objetivo = self._construir_clave_url(url)
        if not clave_objetivo:
            return False

        existentes = model.objects.filter(proyecto=proyecto).values_list("url", flat=True)
        for url_existente in existentes:
            if self._construir_clave_url(url_existente) == clave_objetivo:
                return True
        return False

    def _construir_clave_url(self, url: Optional[str]) -> Optional[str]:
        if not url:
            return None

        normalizada = normalizar_url(url)
        if not normalizada:
            return None

        parsed = urlparse(normalizada)
        netloc = parsed.netloc.lower()
        if netloc.startswith("www."):
            netloc = netloc[4:]
        path = (parsed.path or "").rstrip("/")
        params = parsed.params or ""
        query = parsed.query or ""
        fragment = parsed.fragment or ""
        return "|".join([netloc, path, params, query, fragment]) or None

    def forward_payload(self, endpoint_name: str, payload: Dict[str, Any], headers: Optional[Dict[str, str]] = None):
        headers = headers.copy() if headers else {}

        if "Authorization" not in headers and getattr(self.request, "META", None):
            auth_header = self.request.META.get("HTTP_AUTHORIZATION")
            if auth_header:
                headers["Authorization"] = auth_header

        try:
            relative_url = reverse(endpoint_name)
        except NoReverseMatch:
            logger.error("No se encontró el endpoint '%s' para reenviar el payload", endpoint_name)
            return Response(
                {"detail": f"Endpoint '{endpoint_name}' no encontrado."},
                status=500,
            )

        forward_base_url = getattr(settings, "INGESTION_FORWARD_BASE_URL", None)
        if forward_base_url:
            base_url = forward_base_url.rstrip("/")
            target_url = f"{base_url}{relative_url}"
        elif getattr(self, "request", None) is not None:
            target_url = self.request.build_absolute_uri(relative_url)
        else:
            default_base = getattr(settings, "DEFAULT_DOMAIN", "http://localhost:8000")
            target_url = f"{default_base.rstrip('/')}{relative_url}"

        timeout = getattr(settings, "INGESTION_FORWARD_TIMEOUT", 10)

        try:
            response = requests.post(
                target_url,
                json=payload,
                headers=headers or None,
                timeout=timeout,
            )
        except requests.RequestException as exc:  # pragma: no cover - network failure handling
            logger.error(
                "Error reenviando payload a '%s': %s",
                target_url,
                exc,
            )
            return Response(
                {"detail": "No fue posible reenviar el payload."},
                status=502,
            )

        try:
            content = response.json()
        except ValueError:
            content = {"detail": response.text or "Respuesta vacía"}

        return Response(content, status=response.status_code)
